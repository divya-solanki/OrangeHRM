import openpyxl
from src.library.config_helpers import get_testdata_path


class read_test_data_from_excel:
    def read_excel_file(path, sheet, search_value):
        workbook = openpyxl.load_workbook(path)
        sheet = workbook[sheet]

        records = []
        # list_excel = []
        # for row in sheet.iter_rows(min_row=1, min_col=1):
        #     if row[0].value == search_value:
        #
        #         for i in range(1, len(row)):
        #             item = row[i].value
        #             list_excel.append(item)
        #     # records.append(list_excel)
        # return list_excel

        for row in sheet.iter_rows(min_row=1, min_col=1):
            if row[0].value == search_value:
                list_excel = []
                for i in range(1, len(row)):
                    item = row[i].value
                    list_excel.append(item)

                records.append(list_excel)
        return records
